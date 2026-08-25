"""
XLSX Noise Cleanup Regression Tests
====================================

Proves that the D extraction (noise-only cleanup) correctly removes
NaN and Unnamed artifacts while preserving content.
"""

from __future__ import annotations

import tempfile
import unittest
from pathlib import Path


class TestExcelNoiseCleanup(unittest.TestCase):
    """Regression tests for _clean_excel_text and updated _extract_excel."""

    def _make_xlsx(self, rows: list[list], sheet_name: str = "Sheet1") -> Path:
        """Create a temporary .xlsx file with given rows."""
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name
        for row in rows:
            ws.append(row)
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        tmp.close()
        return Path(tmp.name)

    def test_nan_removed_from_output(self):
        """Extraction does not contain standalone NaN tokens."""
        from kurukshetra.extractors.text_extractor import TextExtractor

        path = self._make_xlsx([
            ["Name", "Value"],
            ["Alpha", 1],
            ["Beta", 2],
        ])
        try:
            text = TextExtractor().extract(path)
            self.assertIsNotNone(text)
            self.assertNotIn("NaN", text)
        finally:
            path.unlink(missing_ok=True)

    def test_unnamed_headers_removed(self):
        """Extraction does not contain 'Unnamed: N' column headers."""
        from kurukshetra.extractors.text_extractor import TextExtractor

        # Create XLSX with merged/empty header cells that cause Unnamed columns
        import openpyxl

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append([None, "Task Subject", "Due Date"])
        ws.append(["Step 1", "Configure G3", "Day 1"])
        ws.append(["Step 2", "Validate", "Day 2"])
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        tmp.close()
        path = Path(tmp.name)

        try:
            text = TextExtractor().extract(path)
            self.assertIsNotNone(text)
            # "Unnamed:" should not appear
            self.assertNotRegex(text, r"Unnamed:\s*\d+")
        finally:
            path.unlink(missing_ok=True)

    def test_sheet_names_preserved(self):
        """Sheet names are preserved in the extraction output."""
        from kurukshetra.extractors.text_extractor import TextExtractor

        import openpyxl

        wb = openpyxl.Workbook()
        wb.active.title = "Workflow"
        wb.active.append(["Step", "Action"])
        wb.active.append(["1", "Install"])
        ws2 = wb.create_sheet("Validation")
        ws2.append(["Check", "Result"])
        ws2.append(["Pass", "OK"])
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        tmp.close()
        path = Path(tmp.name)

        try:
            text = TextExtractor().extract(path)
            self.assertIn("--- Sheet: Workflow ---", text)
            self.assertIn("--- Sheet: Validation ---", text)
        finally:
            path.unlink(missing_ok=True)

    def test_meaningful_content_preserved(self):
        """Meaningful cell content is preserved after cleanup."""
        from kurukshetra.extractors.text_extractor import TextExtractor

        path = self._make_xlsx([
            ["Workflow", "G3 Data Feed Configuration"],
            ["Trigger", "Case Opens"],
            ["Task", "ICS to configure EDF"],
        ])
        try:
            text = TextExtractor().extract(path)
            self.assertIn("G3 Data Feed Configuration", text)
            self.assertIn("Case Opens", text)
            self.assertIn("ICS to configure EDF", text)
        finally:
            path.unlink(missing_ok=True)

    def test_empty_sheet_handled_safely(self):
        """Empty sheets do not cause errors."""
        from kurukshetra.extractors.text_extractor import TextExtractor

        import openpyxl

        wb = openpyxl.Workbook()
        wb.active.title = "Empty"
        # Don't add any data
        tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
        wb.save(tmp.name)
        tmp.close()
        path = Path(tmp.name)

        try:
            text = TextExtractor().extract(path)
            # Should return empty or sheet-only string, no crash
            self.assertIsInstance(text, str)
        finally:
            path.unlink(missing_ok=True)

    def test_clean_spreadsheet_not_damaged(self):
        """A clean spreadsheet without NaN/Unnamed is not damaged by cleanup."""
        from kurukshetra.extractors.text_extractor import TextExtractor

        path = self._make_xlsx([
            ["ID", "Name", "Status"],
            ["1", "G3 RMS", "Active"],
            ["2", "SFDC", "Active"],
            ["3", "Datadog", "Monitoring"],
        ])
        try:
            text = TextExtractor().extract(path)
            self.assertIn("G3 RMS", text)
            self.assertIn("SFDC", text)
            self.assertIn("Datadog", text)
            self.assertIn("Active", text)
            self.assertNotIn("NaN", text)
            self.assertNotIn("Unnamed", text)
        finally:
            path.unlink(missing_ok=True)

    def test_clean_excel_text_helper(self):
        """_clean_excel_text removes NaN and Unnamed, preserves content."""
        from kurukshetra.extractors.text_extractor import TextExtractor

        raw = (
            "Unnamed: 0  Unnamed: 1  Value\n"
            "NaN         Task A      100\n"
            "Step 2      Task B      200\n"
        )
        cleaned = TextExtractor._clean_excel_text(raw)
        self.assertNotIn("Unnamed", cleaned)
        self.assertNotIn("NaN", cleaned)
        self.assertIn("Task A", cleaned)
        self.assertIn("Task B", cleaned)
        self.assertIn("100", cleaned)
        self.assertIn("200", cleaned)

    def test_xls_extraction_applies_same_cleanup(self):
        """XLS extraction also applies NaN/Unnamed cleanup."""
        # Create a .xls file using xlwt if available
        try:
            import xlwt
        except ImportError:
            self.skipTest("xlwt not installed")

        wb = xlwt.Workbook()
        ws = wb.add_sheet("Data")
        ws.write(0, 0, "Name")
        ws.write(0, 1, "Value")
        ws.write(1, 0, "G3 RMS")
        ws.write(1, 1, 100)
        tmp = tempfile.NamedTemporaryFile(suffix=".xls", delete=False)
        wb.save(tmp.name)
        tmp.close()
        path = Path(tmp.name)

        try:
            from kurukshetra.extractors.text_extractor import TextExtractor
            text = TextExtractor().extract(path)
            self.assertIsNotNone(text)
            self.assertIn("G3 RMS", text)
            self.assertIn("100", text)
        finally:
            path.unlink(missing_ok=True)


if __name__ == "__main__":
    unittest.main()
